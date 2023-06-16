from tkinter import *
from PIL import ImageTk, Image
from tkinter import messagebox
import openpyxl
import time

#================================================== TAMPILAN BACKGROUND ====================================================
Accountsystem = Tk()                                                                # variabel utama (Accountsystem)
Accountsystem.state('zoomed')                                                       # agar tampilan layar penuh
Accountsystem.resizable(0,0)                                                        # agar ukuran tidak berubah
Accountsystem.title('KELOMPOK 15')                                                  # judul 

def show_frame(frame):                                                              # fungsi untuk berpindah halaman
    frame.tkraise()
#=================================================== HOMEPAGE & PILIHAN ====================================================
homepage = Frame(Accountsystem)                                                     # tampilan homepage
pilihan = Frame(Accountsystem)                                                      # tampilan halaman ketika memilih random/preferensi

for frame in (homepage, pilihan):                                                   # agar bisa berpindah dari homepage ke halaman pilihan
    frame.grid(row=0,column=0,sticky='nsew')                                        # tata letak halaman

#====================================================== RANDOM PAGE ========================================================
randompage = Frame(Accountsystem,bg='#F6D58E',                                      # tampilan random page
                   width=1500, height=1000)

for frame in (randompage,pilihan):                                                  # agar dapat kembali ke halaman homepage
    frame.grid(row=0,column=0,sticky='nsew')                                        # tata letak homepage

#======================================================== PREFERENSI =======================================================
hlmn = Frame(Accountsystem)                                                         # tampilan preferensi

for frame in (hlmn,pilihan):                                                        # agar bisa berpindah halaman dari homepage ke halaman preferensi
    frame.grid(row=0,column=0,sticky='nsew')                                        # tata letak halaman

#===================================================== LOGIN & BUAT AKUN =====================================================
login = Frame(Accountsystem)                                                        # tampilan login
buat_akun = Frame(Accountsystem)                                                    # tampilan buat akun

for frame in (login, buat_akun):                                                    # agar bisa mengakses login dan buat akun tanpa menutup program terlebih dahulu
    frame.grid(row=0,column=0,sticky='nsew')                                        # tata letak halaman agar otomatis berganti ke halaman login

#============================================================================================================================
#======================================================== NOVITA ARILFA =====================================================
#============================================================================================================================

#============================================================ LOADING =======================================================
latar = Frame(Accountsystem, width=1500, height=1000, bg='#F6D58E').place(x=0,y=0)

food_picker=PhotoImage(file='food picker.png')
fbLabel=Label(Accountsystem,image=food_picker,bg='#F6D58E')
fbLabel.place(x=320,y=150)

Label(latar,text='Loading...',font=('yu gothic ui bold',25,'bold'),bg='#F6D58E',fg='brown').place(x=30,y=670)


b = ImageTk.PhotoImage(Image.open('kosong.png'))
a = ImageTk.PhotoImage(Image.open('ubi.png'))

for i in range(4):
    l1 = Label(latar,image=a, border=0, relief=SUNKEN).place(x=210,y=400)
    l2 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=460,y=400)
    l3 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=710,y=400)
    l4 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=960,y=400)
    Accountsystem.update_idletasks()
    time.sleep(0.5)

    l1 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=210,y=400)
    l2 = Label(latar,image=a, border=0, relief=SUNKEN).place(x=460,y=400)
    l3 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=710,y=400)
    l4 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=960,y=400)
    Accountsystem.update_idletasks()
    time.sleep(0.5)

    l1 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=210,y=400)
    l2 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=460,y=400)
    l3 = Label(latar,image=a, border=0, relief=SUNKEN).place(x=710,y=400)
    l4 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=960,y=400)
    Accountsystem.update_idletasks()
    time.sleep(0.5)

    l1 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=210,y=400)
    l2 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=460,y=400)
    l3 = Label(latar,image=b, border=0, relief=SUNKEN).place(x=710,y=400)
    l4 = Label(latar,image=a, border=0, relief=SUNKEN).place(x=960,y=400)
    Accountsystem.update_idletasks()
    time.sleep(0.5)

show_frame(login)

#===================================================== BUAT AKUN  =========================================================

#================================================== Background Image ======================================================
bg = ImageTk.PhotoImage(file='3.png')
background2 = Label(buat_akun,image=bg)
background2.place(x=0,y=0)

#=================================================== Kumpulan Fungsi ======================================================
def user(event):
    if usernameEntry.get()=='Username':
        usernameEntry.delete(0,END)

def password(event):
    if passwordEntry.get()=='Password':
        passwordEntry.delete(0,END)

def konfirmasi_password_enter(event):
    if konfirmasi_password.get()=='Konfirmasi Password':
        konfirmasi_password.delete(0,END)

def connect_database():
    if usernameEntry.get()=='' or passwordEntry.get()=='' or konfirmasi_password.get()=='':
        messagebox.showwarning('Error','Mohon isi terlebih dahulu')
        wb = openpyxl.load_workbook('registrasi.xlsx')
        sheet = wb.active

    elif usernameEntry.get() == passwordEntry.get():
        messagebox.showwarning('Error','username sama password tidak boleh sama')
        
    elif passwordEntry.get() != konfirmasi_password.get():
        messagebox.showwarning('Error', 'PAssword dan konfirmasi password tidak sama')
    
    elif check.get()==0:
        messagebox.showwarning('Error', 'Mohon isi terms and conditions')

    else:
        username = usernameEntry.get()
        password = passwordEntry.get()
        konfirmasi = konfirmasi_password.get()

        wb = openpyxl.load_workbook('registrasi.xlsx')
        sheet = wb.active

        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = username
        sheet.cell(row=next_row, column=2).value = password

        wb.save('registrasi.xlsx')

        usernameEntry.delete(0, END)
        passwordEntry.delete(0, END)
        konfirmasi_password.delete(0, END)
        
        messagebox.showinfo('Sukses','Akun berhasil dibuat, mohon login terlebih dahulu')
        show_frame(login)

# ========================================================== Tampilan buat akun ================================================
heading=Label(background2,text='BUAT AKUN',font=('yu gothic ui bold',30,'bold'),bg='#FCFDFE',fg='brown')
heading.place(x=560,y=160)

usernameLabel=Label(background2,text='Username',font=('Microsoft Yahei UI Light',10,'bold'),bg='#FCFDFE',fg='brown')
usernameLabel.place(x=500,y=250)

usernameEntry=Entry(background2,width=44,font=('Microsoft Yahei UI Light',10,'bold'),bd=0,fg='#FCFDFE',bg='brown')
usernameEntry.place(x=500,y=280)

passwordLabel=Label(background2,text='Password',font=('Microsoft Yahei UI Light',10,'bold'),bg='#FCFDFE',fg='brown')
passwordLabel.place(x=500,y=310)

passwordEntry=Entry(background2,width=44,font=('Microsoft Yahei UI Light',10,'bold'),bd=0,fg='#FCFDFE',bg='brown')
passwordEntry.place(x=500,y=340)

konfirmasiLabel=Label(background2,text='Konfirmasi Password',font=('Microsoft Yahei UI Light',10,'bold'),bg='#FCFDFE',fg='brown')
konfirmasiLabel.place(x=500,y=370)

konfirmasi_password=Entry(background2,width=44,font=('Microsoft Yahei UI Light',10,'bold'),bd=0,fg='#FCFDFE',bg='brown')
konfirmasi_password.place(x=500,y=400)

check=IntVar()
termsandconditions=Checkbutton(background2,text='I agree to the Terms & Conditions',
                               font=('Microsoft Yahei UI Light',9,'bold'),fg='brown',bg='#FCFDFE',
                               activebackground='#FCFDFE',activeforeground='brown',cursor='hand2',variable=check)
termsandconditions.place(x=500,y=437)

signupButton=Button(background2,text='Buat Akun',font=('Open Sans',15,'bold'),bd=0,bg='brown',
                    fg='#FCFDFE',activebackground='brown',activeforeground='#FCFDFE',width=20,
                    cursor='hand2',command=connect_database)
signupButton.place(x=550,y=480)

signupLabel=Label(background2,text="Sudah punya akun?",font=('Open Sans',9),fg='brown',bg='#FCFDFE')
signupLabel.place(x=590,y=533)

akun_lamaButton=Button(background2,text='Login',font=('Open Sans',9,'bold underline'),
                        fg='brown',bg='#FCFDFE',activeforeground='brown',activebackground='#FCFDFE',
                        cursor='hand2',bd=0,command=lambda:show_frame(login))
akun_lamaButton.place(x=705,y=533)

#============================================================ LOGIN =========================================================
        
# ====================================================== Background Image ===================================================
Login_backgroundImage = ImageTk.PhotoImage(file='3.png')
background = Label(login,image=Login_backgroundImage)
background.place(x=0,y=0)
# ======================================================= Kumpulan fungsi ===================================================
def lupa_password():
    def update():
        lupa_username = email_baru.get()
        lupa_password = password_baru.get()

        wb = openpyxl.load_workbook('registrasi.xlsx')
        sheet = wb.active

        found = False
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == lupa_username:
                found = True
                sheet.cell(row=row, column=2).value = lupa_password
                wb.save('registrasi.xlsx')
                messagebox.showinfo('Sukses','Berhasil mengganti password')
            
        if not found:
            messagebox.showerror('Error','Username tidak ditemukan')

    lupa = Toplevel()
    window_width = 350
    window_height = 290
    screen_width = lupa.winfo_screenwidth()
    screen_height = lupa.winfo_screenheight()
    position_top = int(screen_height / 4 - window_height / 4)
    position_right = int(screen_width / 2 - window_width / 2)
    lupa.geometry(f'{window_width}x{window_height}+{position_right}+{position_top}')

    lupa.title('Lupa Password')
    # win.iconbitmap('images\\aa.ico')
    lupa.configure(background='#F6D58E')
    lupa.resizable(False, False)

    # =================== Email ====================
    email_baru = Entry(lupa, bg='#FCFDFE', font=("yu gothic ui semibold", 12),bd=0,fg='brown')
    email_baru.place(x=50, y=60, width=256, height=35)
    email_label = Label(lupa, text='Username', fg="brown", bg='#F6D58E',font=("yu gothic ui", 11, 'bold'))
    email_label.place(x=50, y=30)

    # ============== Password Baru ==================
    password_baru = Entry(lupa, bg="#FCFDFE", font=("yu gothic ui semibold", 12), show='•',bd=0,fg='brown')
    password_baru.place(x=50, y=130, width=256, height=35)
    password_label = Label(lupa, text='Password Baru', fg="brown", bg='#F6D58E',font=("yu gothic ui", 11, 'bold'))
    password_label.place(x=50, y=100)

    # ============== Update password Button ====================
    update_pass = Button(lupa, fg='#FCFDFE', text='Update Password', bg='brown', font=("yu gothic ui", 12, "bold"),
                         cursor='hand2', relief="flat", bd=0, highlightthickness=0, activebackground="brown",command=lambda:update())
    update_pass.place(x=50, y=200, width=256, height=45)
    
def hide():
    openeye.config(file='closeye.png')
    password.config(show='*')
    eyeButton.config(command=show)

def show():
    openeye.config(file='openeye.png')
    password.config(show='')
    eyeButton.config(command=hide)

def user_enter(event):
    if username.get()=='Username':
        username.delete(0,END)

def password_enter(event):
    if password.get()=='Password':
        password.delete(0,END)

def masuk ():
    login_username = username.get()
    login_password = password.get()

    wb = openpyxl.load_workbook('registrasi.xlsx')
    sheet = wb.active

    found = False
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == login_username and sheet.cell(row=row, column=2).value == login_password:
            found = True
            break

    if username.get()=='Username' or password.get()=='Password':
        messagebox.showerror('Error', 'Isi dulu yaa')    
    elif found:
        username.delete(0,END)
        password.delete(0,END)

        show_frame(homepage)

    else:
        messagebox.showerror('Error','username/password salah')
# ===================================================== Tampilan login ========================================================
heading=Label(background,text='LOGIN',font=('yu gothic ui bold',30,'bold'),bg='#FCFDFE',fg='brown')
heading.place(x=610,y=160)

username=Entry(background,width=25,font=('Microsoft Yeahei UI Light',11,'bold'),bd=0,bg='#FCFDFE',fg='brown')
username.place(x=500,y=280)
username.insert(0,'Username')
username.bind('<FocusIn>',user_enter)

frame1=Frame(background,width=350,height=2,bg='brown').place(x=500,y=305)

password=Entry(background,width=25,font=('Microsoft Yeahei UI Light',11,'bold'),bd=0,bg='#FCFDFE',fg='brown')
password.place(x=500,y=350)
password.insert(0,'Password')
password.bind('<FocusIn>',password_enter)

frame2 = Frame(background,width=350,height=2,bg='brown').place(x=500,y=375)

openeye = PhotoImage(file='openeye.png')
eyeButton = Button(background,image=openeye,bd=0,bg='#FCFDFE',activebackground='#FCFDFE',
                   cursor='hand2',command=hide)
eyeButton.place(x=820,y=347)

forgetButton=Button(background,text='Lupa Password?',bd=0,bg='#FCFDFE',activebackground='#FCFDFE',
                    cursor='hand2',font=('Microsoft Yeahei UI Light',9,'bold'),fg='brown',
                    activeforeground='brown',command=lambda : lupa_password())
forgetButton.place(x=750,y=385)

loginButton=Button(background,text='Login',font=('Open Sans',15,'bold'),fg='#FCFDFE',
                   bg='brown',activeforeground='#FCFDFE',activebackground='brown',cursor='hand2',
                   bd=0,width=20,command=masuk)

loginButton.place(x=550,y=480)

signupLabel=Label(background,text="Belum punya akun?",font=('Open Sans',9),fg='brown',bg='#FCFDFE')
signupLabel.place(x=578,y=533)

newaccountButton=Button(background,text='Buat akun',font=('Open Sans',9,'bold underline'),
                        fg='brown',bg='#FCFDFE',activeforeground='brown',activebackground='#FCFDFE',
                        cursor='hand2',bd=0, command=lambda : show_frame(buat_akun))
newaccountButton.place(x=693,y=533)

#===================================================================================================================
#====================================================== ILHAM RYAN =================================================
#===================================================================================================================

#======================================================= HOMEPAGE ==================================================
bg3 = ImageTk.PhotoImage(file='homepage.png')
background3 = Label(homepage,image=bg3)
background3.place(x=0,y=0)

lanjut = Button(background3,text='LANJUT',font=('Segoe Print',17,'bold'),bd=0,bg='#82301A',
                    fg='#FCFDFE',cursor='hand2',activebackground='#F6D58E',activeforeground='#82301A',width=12,height=1,
                    command=lambda:show_frame(pilihan))
lanjut.place(x=1115,y=600)

#======================================================= Pilihan ===================================================
bg4 = ImageTk.PhotoImage(file='pilihan.png')
background4 = Label(pilihan,image=bg4)
background4.place(x=0,y=0)

random = Button(background4,text='RANDOM',font=('Segoe Print',17,'bold'),bd=0,bg='#82301A',
                    fg='#FCFDFE',cursor='hand2',activebackground='#FCFDFE',
                    activeforeground='#82301A',width=17,height=1,command=lambda:show_frame(randompage))
random.place(x=625,y=280)

preferensi = Button(background4,text='PREFERENSI',font=('Segoe Print',17,'bold'),bd=0,bg='#82301A',
                    fg='#FCFDFE',cursor='hand2',activebackground='#FCFDFE',
                    activeforeground='#82301A',width=17,height=1,command=lambda:show_frame(hlmn))
preferensi.place(x=960,y=280)

Accountsystem.mainloop()